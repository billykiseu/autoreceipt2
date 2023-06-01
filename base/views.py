from django.shortcuts import render, redirect
from .models import CustomUser, category, workbook, Profile, Receipt
from django.contrib.auth.decorators import login_required
from django.contrib.messages.views import SuccessMessageMixin
from django.contrib.auth.views import PasswordResetView, PasswordChangeView
from django.contrib.admin import AdminSite
from django.contrib.auth import authenticate, login, logout
from django.urls import reverse_lazy
from .forms import SignUpX, RecoverX
from django.utils.encoding import force_bytes
from django.utils.http import urlsafe_base64_encode
from django.contrib.auth.tokens import default_token_generator
from django.http import JsonResponse, HttpResponseBadRequest, Http404, HttpResponse, FileResponse
from django.core.mail import send_mail
from django.template.loader import render_to_string
from .forms import CustomPasswordResetForm
import os
from django.conf import settings
from django.db.models import Q

#libraries
from openpyxl import load_workbook
from io import BytesIO
from django.shortcuts import render, redirect, HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from bs4 import BeautifulSoup
from datetime import datetime
import os
import tempfile
import zipfile
import shutil
from django.core.mail import EmailMessage
from django.contrib import messages

#savingexcel to db
from django.db import IntegrityError
from django.utils.crypto import get_random_string
from decimal import Decimal, InvalidOperation
from django.db import transaction
from django.core.exceptions import ValidationError

#reciptsdb
from django.core.paginator import Paginator
import math
from django.contrib.staticfiles.storage import staticfiles_storage

#Customadminheader
class MyAdminSite(AdminSite):
    site_header = 'Autorecpt2'
    
#profile
class ChangePasswordView(SuccessMessageMixin, PasswordChangeView):
    template_name = 'changepassword.html'
    success_message = "Successfully Changed Your Password"
    success_url = reverse_lazy('dashboard')

#Passwordreset
class CustomPasswordResetView(PasswordResetView):
    form_class = CustomPasswordResetForm
    success_url = '/password-reset-done/'
    template_name = 'reset.html'

    def password_reset(request):
        if request.method == 'POST':
            form = CustomPasswordResetForm(request.POST)
            
        else:
            form = CustomPasswordResetForm()

        return render(request, 'password_reset.html', {'form': form})
     
# Credentials-Stuff
def loginfail(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']

        CustomUser = authenticate(request, username=username, password=password)
        if CustomUser is not None:
            login(request, CustomUser)
            #user_logged_in.send(user, instance=user, request=request)
            return redirect('dashboard')
        else:
            return redirect('loginfail')

    return render(request, 'loginfail.html')

def loginUser(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']

        CustomUser = authenticate(request, username=username, password=password)
        if CustomUser is not None:
            login(request, CustomUser)
            return redirect('dashboard')
        else:
            return redirect('loginfail')

    return render(request, 'login.html')

def logoutUser(request):
    logout(request)
    return redirect('home')

def signupUser(request):
    form = SignUpX()
    if request.method == "POST":
        form = SignUpX(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.set_password(form.cleaned_data.get("password1"))
            user.save()
            username = form.cleaned_data.get("username")
            password = form.cleaned_data.get("password1")
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect("dashboard")
            else:
                return redirect("signup")

    context = {'form': form}
    return render(request, 'signup.html', context)

def recover(request):
    recover_form = RecoverX()
    if request.method == "POST":
        recover_form = RecoverX(request.POST)
        if recover_form.is_valid():
            data = recover_form.cleaned_data.get('email')
            user_email = CustomUser.objects.filter(email=data).first()
            if user_email is not None:
                subject = 'Password Reset request'
                email_template_name = 'recovertemplate.txt'
                parameters = {
                    'email': user_email.email,
                    'domain': '127.0.0.1:5000',
                    'site_name': 'XendMe.com',
                    'uid': urlsafe_base64_encode(force_bytes(user_email.pk)),
                    'token': default_token_generator.make_token(user_email),
                    'protocol': 'http',
                }
                email = render_to_string(email_template_name, parameters)
                try:
                    send_mail(subject, email, '', [
                              user_email.email], fail_silently=False)
                except:
                    return HttpResponse('Invalid Header')
                return redirect('recoverdone')
            else:
                return redirect('recoverfail')

    context = {
        'recover_form': recover_form,
    }
    return render(request, 'recover.html', context)

def recoverfail(request):
    recover_form = RecoverX()
    if request.method == "POST":
        recover_form = RecoverX(request.POST)
        if recover_form.is_valid():
            data = recover_form.cleaned_data.get('email')
            user_email = CustomUser.objects.filter(email=data).first()
        if user_email is not None:
            subject = 'Password Reset request'
            email_template_name = 'base/recovertemplate.txt'
            parameters = {
                'email': user_email.email,
                'domain': '127.0.0.1:5000',
                'site_name': 'Billykiseu.com',
                'uid': urlsafe_base64_encode(force_bytes(user_email.pk)),
                'token': default_token_generator.make_token(user_email),
                'protocol': 'http',
            }
            email = render_to_string(email_template_name, parameters)
            try:
                send_mail(subject, email, '', [
                          user_email.email], fail_silently=False)
            except:
                return HttpResponse('Invalid Header')
            return redirect('recoverdone')
        else:
            return redirect('recoverfail')

    context = {
        'recover_form': recover_form,
    }
    return render(request, 'recoverfail.html', context)

#homepage
def home(request):
    context = {
        'authenticated': request.user.is_authenticated,
    }
    return render(request, 'home.html', context)

#dashbaord, read Excel file
@login_required
def dashboard(request):
    if request.method == 'POST' and request.FILES.get('file'):
        uploaded_file = request.FILES['file']
        
        # Check if the file type is valid
        allowed_file_types = ['.xlsx', '.xls']
        file_extension = os.path.splitext(uploaded_file.name)[1]
        if file_extension not in allowed_file_types:
            error_message = "Invalid file type. Please upload an Excel file (.xlsx or .xls)."
            messages.error(request, error_message)
            return redirect('dashboard')

        # Check if a workbook with the same name already exists
        try:
            existing_workbook = workbook.objects.get(name=uploaded_file.name)
            # Delete the existing file associated with the workbook
            existing_workbook.file.delete(save=False)
        except workbook.DoesNotExist:
            existing_workbook = None
        # Create a new instance of the workbook model or use the existing one
        new_workbook = existing_workbook or workbook()

        # Set the file and name fields of the workbook
        new_workbook.file = uploaded_file
        new_workbook.name = uploaded_file.name
        new_workbook.save()
        
        # Load the workbook using openpyxl
        workbook_obj = load_workbook(uploaded_file)
        sheet = workbook_obj.active
        transactions = []
        header_row = [cell.value for cell in sheet[1]]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = []
            for i, cell in enumerate(row):
                header = header_row[i] if i < len(header_row) else ''
                if isinstance(cell, datetime):
                    row_data.append(cell.strftime('%Y-%m-%d'))  # Convert datetime to string
                elif header == 'Date':
                    # Try to infer the date format
                    valid_date = False
                    for date_format in ['%d/%m/%Y', '%d.%m.%Y']:
                        try:
                            date_obj = datetime.strptime(str(cell), date_format)
                            row_data.append(date_obj.strftime('%Y-%m-%d'))
                            valid_date = True
                            break
                        except ValueError:
                            pass
                    if not valid_date:
                        row_data.append('')  # Invalid date format
                else:
                    row_data.append(cell)
            transaction = dict(zip(header_row, row_data))
            transactions.append(transaction)
        
        # Save the transactions in the session and pass them to the preview.html template
        request.session['transactions'] = transactions
        return render(request, 'preview.html', {'transactions': transactions})
    
    recent_workbooks = workbook.objects.all().order_by('-updated')[:10]
    context = {
        'recent_workbooks': recent_workbooks,
        'recent_count': recent_workbooks.count(),
    }
 
    return render(request, 'dashboard.html', context)

#save excel data to db
def generate_receipt_number():
    prefix = 'EDVL'
    random_string = get_random_string(length=6)
    return f'{prefix}-{random_string}'

def infer_date_format(date_str):
    try:
        date_obj = datetime.strptime(str(date_str), '%Y-%m-%d')
        return date_obj.strftime('%Y-%m-%d')
    except ValueError:
        return None
    
def savetodb(request):
    transactions = request.session.get('transactions', [])
    error_records = []

    with transaction.atomic():
        for transaction_data in transactions:
            try:
                receipt = Receipt()

                receipt_number = generate_receipt_number()
                receipt.receipt_number = receipt_number

                receipt.house_number = transaction_data.get('House Number')
                receipt.name = transaction_data.get('Name')

                # Infer the date format
                transaction_date = infer_date_format(transaction_data.get('Date'))
                if transaction_date is None:
                    error_records.append({
                        'transaction': transaction_data,
                        'error_message': f"Invalid date format for record: {transaction_data}."
                    })
                    continue

                receipt.date = transaction_date
                receipt.phone = transaction_data.get('Phone')
                receipt.description = transaction_data.get('Description')

                try:
                    amount = Decimal(str(transaction_data.get('Amount')).replace(',', ''))
                    receipt.amount = amount
                except (ValueError, InvalidOperation):
                    error_records.append({
                        'transaction': transaction_data,
                        'error_message': f"Invalid amount format for record: {transaction_data}."
                    })
                    continue

                receipt.email = transaction_data.get('Email')
                receipt.pending = False
                receipt.save()
            except IntegrityError:
                error_records.append({
                    'transaction': transaction_data,
                    'error_message': f"Receipt number already exists for record: {transaction_data}."
                })
            except Exception as e:
                error_records.append({
                    'transaction': transaction_data,
                    'error_message': f"Error saving record: {transaction_data}. {str(e)}"
                })

    success_count = len(transactions) - len(error_records)
    error_count = len(error_records)

    success_message = f"{success_count} records saved successfully."
    messages.success(request, success_message)

    if error_records:
        request.session['error_records'] = error_records

        error_message = f"{error_count} records encountered errors."
        messages.error(request, error_message)

        return redirect('edit_records')

    return redirect('dashboard')

def edit_records(request):
    error_records = request.session.get('error_records', [])

    if request.method == 'POST':
        transaction_id = request.POST.get('transaction_id')
        transaction = next((rec['transaction'] for rec in error_records if rec['transaction'].get('id') == transaction_id), None)

        if transaction:
            # Update the transaction with the edited field(s)
            field_to_edit = request.POST.get('field_to_edit')  # Get the field to be edited from the form
            new_value = request.POST.get('new_value')  # Get the new value entered by the user

            if field_to_edit and field_to_edit in transaction['transaction']:
                # Update the specified field with the new value
                transaction['transaction'][field_to_edit] = new_value

                # Save the updated transaction
                try:
                    if 'error_message' not in transaction:
                        # Only save the updated transaction if it's not a date error
                        receipt = Receipt()
                        receipt_number = generate_receipt_number()
                        receipt.receipt_number = receipt_number

                        for field, value in transaction['transaction'].items():
                            setattr(receipt, field, value)

                        receipt.save()

                    # Remove the updated transaction from error_records
                    error_records = [rec for rec in error_records if rec['transaction'].get('id') != transaction_id]
                    request.session['error_records'] = error_records
                except Exception as e:
                    error_message = f"Error saving record: {transaction}. {str(e)}"
                    messages.error(request, error_message)
            else:
                error_message = "Invalid field or field not found in the transaction."
                messages.error(request, error_message)

        # If there are no more error records or only date errors, redirect to the dashboard
        if not error_records or all('error_message' in rec for rec in error_records):
            return redirect('dashboard')

    context = {
        'error_records': error_records
    }
    return render(request, 'savetodb.html', context)

#receipttemplate
def template(request):
    return render(request, 'receipt_template.html')

#Receipt
@login_required
def receipts(request):
    receipts = Receipt.objects.all()
    paginator = Paginator(receipts, 20)  # Display 40 receipts per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {'page_obj': page_obj, 
               'receipts_count': receipts.count()}
    return render(request, 'receipt.html', context)

#search
@login_required
def search(request):
    q = request.GET.get('q') if request.GET.get('q') is not None else ''
    search = Receipt.objects.filter(
        Q(receipt_number__icontains=q) |
        Q(house_number__icontains=q) |
        Q(name__icontains=q) |
        Q(description__icontains=q) |
        Q(email__icontains=q)
    ).order_by('-date')

    paginator = Paginator(search, 10)  # Show 10 items per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'search': page_obj,
        'item_count': search.count(),
    }
    return render(request, 'searchresults.html', context)

#download manual
def manual(request):
    file_path = os.path.join(settings.MEDIA_ROOT, 'manual.pdf')
    return FileResponse(open(file_path, 'rb'), as_attachment=True, filename='manual.pdf')

#gen and download
def generate_receipts(request):
    # Fetch all the records from the Receipt model
    receipts = Receipt.objects.all()

    # Create a temporary in-memory zip file
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
        # Group receipts by house number
        grouped_receipts = {}
        for receipt in receipts:
            house_number = receipt.house_number
            if house_number not in grouped_receipts:
                grouped_receipts[house_number] = []
            grouped_receipts[house_number].append(receipt)

        # Generate PDFs and organize in subdirectories
        for house_number, receipts in grouped_receipts.items():
            # Create a subdirectory for the house number
            subdirectory = f"{house_number}/"
            zip_file.writestr(subdirectory, '')

            # Generate and add PDFs for each receipt in the subdirectory
            for receipt in receipts:
                # Generate the receipt filename
                filename = f"{receipt.receipt_number}.pdf"

                # Render the HTML template for the receipt
                html_content = render_to_string('receipt_template.html', {'receipt': receipt})

                # Generate PDF from HTML content
                pdf_content = generate_pdf(html_content)

                # Add the PDF content to the subdirectory within the ZIP file
                zip_file.writestr(os.path.join(subdirectory, filename), pdf_content)

    # Seek to the beginning of the in-memory zip file
    zip_buffer.seek(0)

    # Create an HttpResponse with the ZIP file contents
    response = HttpResponse(zip_buffer, content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename="receipts.zip"'
    return response

def generate_pdf(html_content):
    # Replace static file URLs with absolute URLs
    absolute_url = staticfiles_storage.url('')
    html_content = html_content.replace('{% static "', f'{absolute_url}')
    html_content = html_content.replace('{% static "', f'{absolute_url}')
    
    # Create a BytesIO buffer to receive the generated PDF file
    pdf_buffer = BytesIO()

    # Generate PDF using the HTML content
    pisa.CreatePDF(html_content, dest=pdf_buffer)

    # Seek to the beginning of the PDF buffer
    pdf_buffer.seek(0)

    # Retrieve the PDF content
    pdf_content = pdf_buffer.getvalue()
    pdf_buffer.close()

    return pdf_content

#gen and send
def send_receipts(request):
    # Fetch all the unique house numbers
    house_numbers = Receipt.objects.values_list('house_number', flat=True).distinct()
    total_houses = len(house_numbers)
    emails_sent = 0
    emails_failed = 0

    # Iterate over each house number
    for house_number in house_numbers:
        # Fetch all the receipts for the current house number
        receipts = Receipt.objects.filter(house_number=house_number)

        # Create a list to store email addresses for the current house number
        email_addresses = []

        # Iterate over each receipt
        for receipt in receipts:
            # Render the HTML template for the receipt
            html_content = render_to_string('receipt_template.html', {'receipt': receipt})

            # Generate PDF from HTML content
            pdf_content = generate_pdf(html_content)

            # Create a temporary file to save the PDF
            pdf_path = f"{receipt.receipt_number}.pdf"
            with open(pdf_path, 'wb') as pdf_file:
                pdf_file.write(pdf_content)

            # Add the email address to the list
            email_addresses.append(receipt.email)

        # Compose email message
        subject = "Receipts for your purchases"
        message = "Dear customer,\n\nPlease find attached the receipts for your recent purchases.\n\nThank you for your business!"
        from_email = settings.DEFAULT_FROM_EMAIL
        to_email = email_addresses

        # Create email attachments for each PDF
        attachments = []
        for receipt in receipts:
            pdf_path = f"{receipt.receipt_number}.pdf"
            with open(pdf_path, 'rb') as pdf_file:
                attachments.append((os.path.basename(pdf_path), pdf_file.read(), 'application/pdf'))

            # Delete the temporary PDF file
            os.remove(pdf_path)

        # Send email with attachments
        email = EmailMessage(subject, message, from_email, to_email)
        email.attachments = attachments

        try:
            email.send()
            # Update the count of sent emails
            emails_sent += 1
        except Exception:
            # Update the count of failed emails
            emails_failed += 1

    # Redirect to the receipts page with a success message
    messages.success(request, f"{emails_sent} emails sent successfully. {emails_failed} emails failed.")
    return redirect('receipts')


#clean
@login_required
def clean(request):
    if request.method == 'POST':
        file = request.FILES.get('file')

        # Check if a file is uploaded
        if file:
            # Get the file extension
            file_extension = file.name.split('.')[-1].lower()

            # Define the allowed file extensions
            allowed_extensions = ['qfx', 'ofx', 'pdf', 'csv']

            # Validate the file extension
            if file_extension not in allowed_extensions:
                error_message = "Invalid file type. Please upload a .qfx, .ofx, .pdf, or .csv file."
                messages.error(request, error_message)
                return redirect('clean')

            # Check if a workbook with the same name already exists
            try:
                existing_workbook = workbook.objects.get(name=file.name)
                # Delete the existing file associated with the workbook
                existing_workbook.file.delete(save=False)
            except workbook.DoesNotExist:
                existing_workbook = None

            # Create a new instance of the workbook model or use the existing one
            new_workbook = existing_workbook or workbook()

            # Set the file and name fields of the workbook
            new_workbook.file = file
            new_workbook.name = file.name

            try:
                # Validate the workbook instance
                new_workbook.full_clean()
                new_workbook.save()

                # File is valid and saved successfully
                return HttpResponse('File uploaded and processed successfully.')
            except ValidationError as e:
                # File validation failed
                error_message = str(e)
                messages.error(request, error_message)
                return redirect('clean')

    return render(request, 'clean.html')

            
            
           